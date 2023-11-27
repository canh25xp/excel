from openpyxl import Workbook, workbook

wb = Workbook()

ws = wb.active

ws['A1'] = 'Hello Rachel'
ws['A2'] = 5702

wb.save('sample.xlsx')