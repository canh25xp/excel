from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Series, Reference

wb = load_workbook(filename='sample.xlsx')

ws = wb.active

cols = [col for col in ws.iter_cols(min_row=2, max_col=4, max_row=8, values_only = True)]

names = list(cols[0])
dates = list(cols[1])
money = list(cols[2])
docter = list(cols[3])

chart = BarChart()
chart.type = "col"
chart.title = "Daily report"
chart.y_axis.title = "money"
chart.x_axis.title = "day"

data = Reference(ws, 3, 2, 3, 8)
categories = Reference(ws, 1, 2, 1, 8)

chart.add_data(data)
chart.set_categories(categories)

ws_chart = wb.create_sheet("chart") 
ws_chart.add_chart(chart, "A1")

wb.save('out.xlsx')